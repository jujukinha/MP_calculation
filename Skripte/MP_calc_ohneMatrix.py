from collections import defaultdict
import pandas as pd
from glob import glob
import numpy as np
import os
import itertools
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from MP_final_Funktionen import read_data, ratio_PE, save_excel
pd.options.mode.chained_assignment = None #damit keine Warning message kommt
from TGA_final import run_tga

###____________________Skript für externe Kalibrierung ohne Matrix vom Juni, November 2020 und März 2021___________###
##_____________________Skript für Standards____________________________________##

# Einlesen von Kalibrierungsdaten aus Excel "Kali-Daten" für Bewertung
## In der Excel sind von allen 3 Kalibrierungen die Werte gespeichert

v_dev_june = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='ratio_june', index_col='Name')
v_dev_nov = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='ratio_nov', index_col='Name')
v_dev_mar = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='ratio_mar', index_col='Name')
v_dev_aug = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='ratio_aug', index_col='Name')

df_benchmark_june = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='benchmark_june', index_col='Proben-ID')
df_benchmark_nov = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='benchmark_nov', index_col='Proben-ID')
df_benchmark_mar = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='benchmark_mar', index_col='Proben-ID')
df_benchmark_mar_newRT = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='benchmark_mar_newRT', index_col='Proben-ID')
df_benchmark_aug = pd.read_excel(r'Kalibrierungsdaten\Kali_Daten.xlsx', sheet_name='benchmark_aug', index_col='Proben-ID')

# Proben-Daten einlesen für Bewertung ohne BW-file
## Dafür detailed reports in Ordner "Proben_Rohdaten" packen

filenames_samples = glob('Proben_Rohdaten\\*.txt')

filenames_samples_rating = [ x for x in filenames_samples if not x.endswith('BW.txt')]
open_samples = [open(f, 'r+') for f in filenames_samples_rating]

sample_list = read_data(open_samples)

filenames_samples_rating = [s.replace('Proben_Rohdaten\\', '') for s in filenames_samples_rating]

ratio_list = sample_list

# Benchmark-system

gut = 1
mittelgut = 0.75
mittel = 0.5
mittelschlecht = 0.25
schlecht = 0

# Funktion für Verhältnis PE 55/81 aufrufen

df_ratio = ratio_PE(ratio_list)

# Datumsabgleich für richtige file-Auswahl
## Wenn neue Kali dazukommt hier ergänzen

date_nov = '201116'
date_kali_nov  = datetime.strptime(date_nov, '%y%m%d')
date_mar = '210305'
date_kali_mar  = datetime.strptime(date_mar, '%y%m%d')
date_may = '210419'
date_kali_may  = datetime.strptime(date_may, '%y%m%d')
date_aug = '210823'
date_kali_aug = datetime.strptime(date_aug, '%y%m%d')


# Bewertung Verhältnis 55/81
## Ergebnisse werden in einer Liste gespeichert

choices = [gut, mittel, schlecht]

ratio_bm = []
for name, dataframe in zip(filenames_samples_rating, df_ratio):
    date = name[:6]
    date_sample = datetime.strptime(date, '%y%m%d')
    if date_sample < date_kali_nov:
        v_dev = v_dev_june
    if date_kali_nov < date_sample < date_kali_mar:
        v_dev = v_dev_nov
    if date_kali_mar < date_sample < date_kali_aug:
        v_dev = v_dev_mar
    else:
        v_dev = v_dev_aug
    conditions = [(dataframe['Response'] <= v_dev['ratio_mean'] + v_dev['ratio_dev']) &  # gut
                  (v_dev['ratio_mean'] - v_dev['ratio_dev'] <= dataframe['Response']),  # gut
                  (v_dev['ratio_min'] <= dataframe['Response']) &  # mittel
                  (dataframe['Response'] <= v_dev['ratio_max']),  # mittel
                  (dataframe['Response'] < v_dev['ratio_min']) |  # schlecht
                  (v_dev['ratio_max'] < dataframe['Response'])]  # schlecht
    dataframe['ratio_BM'] = np.select(conditions, choices, np.nan)

    ratio_mtw = dataframe['ratio_BM'].mean(axis=0)  # auskommentieren, wenn man alle Werte einzeln sehen möchte
    ratio_mtw2 = round(ratio_mtw, 1)
    ratio_bm.append(ratio_mtw2)

# print(ratio_bm)

# Bewertung RT und Qualifier
## Erstellung dataframes

df_samples_list = []
for dicts in sample_list:
    sample_dict = dict(dicts)
    sample_df = pd.DataFrame.from_dict(sample_dict,
                                 orient='index', columns=['RT', 'RT_start', 'Response', 'Quali1', 'Quali2', 'Quali3'])
    sample_df['Quali1'] = sample_df['Quali1'].replace(0.0, np.nan)
    sample_df['Quali2'] = sample_df['Quali2'].replace(0.0, np.nan)
    sample_df['Quali3'] = sample_df['Quali3'].replace(0.0, np.nan)
    df_samples_list.append(sample_df)

# leere Listen für Bewertungswerte
rt = []
quali = []

# Listen für for-Schleifen
qualis = ['Quali1', 'Quali2', 'Quali3']
qualis_mean = ['Quali1_mean', 'Quali2_mean', 'Quali3_mean']
qualis_dev = ['Quali1_dev', 'Quali2_dev', 'Quali3_dev']
qualis_max = ['Quali1_max', 'Quali2_max', 'Quali3_max']
qualis_min = ['Quali1_min', 'Quali2_min', 'Quali3_min']

# Speicherung der Kalibrierungsdaten
for name in (filenames_samples_rating):
    date = name[:6]
    date_sample = datetime.strptime(date, '%y%m%d')
    if date_sample <= date_kali_nov:
        df_benchmark = df_benchmark_june
    elif date_kali_nov <= date_sample < date_kali_mar:
        df_benchmark = df_benchmark_nov
    elif date_kali_mar <= date_sample < date_kali_may:
        df_benchmark = df_benchmark_mar
    elif date_kali_may <= date_sample < date_kali_aug:
        df_benchmark = df_benchmark_mar_newRT
    else:
        df_benchmark = df_benchmark_aug

    kali_PE = df_benchmark.iloc[4:16, :]
    kali_PP = df_benchmark.iloc[27:33, :]
    kali_PS = df_benchmark.iloc[24:27, :]
    kali_PET1 = df_benchmark.iloc[16:17, :]
    kali_PET2 = df_benchmark.iloc[18:19, :]
    kali_PET3 = df_benchmark.iloc[20:21, :]
    kali_PET4 = df_benchmark.iloc[22:24, :]
    PET_list = [kali_PET1, kali_PET2, kali_PET3, kali_PET4]
    kali_PET = pd.concat(PET_list)
    kali_PMMA = df_benchmark.iloc[0:1, :]
    if date_sample >= date_kali_aug:
        kali_PA = df_benchmark.iloc[33:34, :]
        kali_SBR = df_benchmark.iloc[35:36, :]


#df_polymer_kalis = [kali_PE, kali_PP, kali_PS, kali_PET, kali_PMMA]
#df_polymer_kalis_new = [kali_PE, kali_PP, kali_PS, kali_PET, kali_PMMA, kali_PA, kali_SBR]

# Mega for-Schleife zur Ermittlung der Bewertung
for name, dataframe in zip(filenames_samples_rating, df_samples_list):
    date = name[:6]
    date_sample = datetime.strptime(date, '%y%m%d')
    #Bewertung von RT-Zeit
    df_PE = dataframe.iloc[4:16,:]
    df_PP = dataframe.iloc[27:33,:]
    df_PS = dataframe.iloc[24:27,:]
    df_PET1 = dataframe.iloc[16:17,:]
    df_PET2 = dataframe.iloc[18:19,:]
    df_PET3 = dataframe.iloc[20:21,:]
    df_PET4 = dataframe.iloc[22:24,:]
    PET_list = [df_PET1, df_PET2, df_PET3, df_PET4]
    df_PET = pd.concat(PET_list)
    df_PMMA = dataframe.iloc[0:1,:]
    if date_sample <= date_kali_aug:
        None
    elif date_sample >= date_kali_aug:
        df_PA = dataframe.iloc[33:34, :]
        df_SBR = dataframe.iloc[35:36, :]

    #df_polymer_samples = [df_PE, df_PP, df_PS, df_PET, df_PMMA]
    #df_polymer_samples_new = [df_PE, df_PP, df_PS, df_PET, df_PMMA, df_PA, df_SBR]

    # TODO: hier überlegen wie PA und SBR am besten integriert werden
    if date_sample <= date_kali_aug:
        df_polymer_samples = [df_PE, df_PP, df_PS, df_PET, df_PMMA]
        df_polymer_kalis = [kali_PE, kali_PP, kali_PS, kali_PET, kali_PMMA]
    else:
        df_polymer_samples = [df_PE, df_PP, df_PS, df_PET, df_PMMA, df_PA, df_SBR]
        df_polymer_kalis = [kali_PE, kali_PP, kali_PS, kali_PET, kali_PMMA, kali_PA, kali_SBR]

    for df_samples, df_kalis in zip(df_polymer_samples, df_polymer_kalis):
        condi_rt = [(df_samples['RT'] <= df_kalis['RT_mean'] + df_kalis['RT_dev']) &
                    (df_kalis['RT_mean'] - df_kalis['RT_dev'] <= df_samples['RT']),
                    (df_kalis['RTstart_mean'] <= df_samples['RT']) &
                    (df_samples['RT'] <= df_kalis['RTstart_mean'] + 1),
                    (df_samples['RT'] < df_kalis['RTstart_mean']) |
                    (df_kalis['RTstart_mean'] + 1 < df_samples['RT'])]
        df_samples['RT_bm'] = np.select(condi_rt, choices, np.NaN)
        rt_mean = df_samples['RT_bm'].mean(axis=0)
        rt_mean = round(rt_mean, 1)
        #print(df_samples, rt_mean)
        rt.append(rt_mean)

        #Bewertung von Qualis
        for q, mean, d, ma, mi in zip(qualis, qualis_mean, qualis_dev, qualis_max, qualis_min):
            condi = [(df_samples[q] <= df_kalis[mean] + df_kalis[d]) &   #gut
                     (df_kalis[mean] - df_kalis[d] <= df_samples[q]),
                     (df_kalis[mi] <= df_samples[q]) &                   #mittel
                     (df_samples[q] <= df_kalis[ma]),
                     (df_samples[q] < df_kalis[mi]) |                    #schlecht
                     (df_kalis[ma] < df_samples[q])]
            df_samples['Quali_bm'] = np.select(condi, choices, np.nan)
            #print(df_samples['Quali_bm'])
            #printen wenn man alle Werten einzeln sehen will
            q_mean = df_samples['Quali_bm'].mean(axis=0)
            q_mean = round(q_mean, 1)
            quali.append(q_mean)

# Werte in richtige Reihenfolge sortieren
dict_PE = defaultdict(list)
dict_PP = defaultdict(list)
dict_PS = defaultdict(list)
dict_PET = defaultdict(list)
dict_PMMA = defaultdict(list)
dict_PA = defaultdict(list)
dict_SBR = defaultdict(list)

# if not dict_PA == None:
#     dict_PA = defaultdict(list)
#     dict_SBR = defaultdict(list)

rt_value = int(len(rt) / len(df_samples_list))
quali_value = int(len(quali) / len(df_samples_list))

loop_rt = range(0, int(len(rt)), rt_value)
loop_q = range(0, int(len(quali)), quali_value)


for i, j, ratio, names in zip(loop_rt, loop_q, ratio_bm, filenames_samples_rating):
    date = names[:6]
    date_sample = datetime.strptime(date, '%y%m%d')

    part_rt = rt[i:i + rt_value]
    part_q = quali[j:j + quali_value]

    # RT
    dict_PE[names[:-4]].append(part_rt[0])
    dict_PP[names[:-4]].append(part_rt[1])
    dict_PS[names[:-4]].append(part_rt[2])
    dict_PET[names[:-4]].append(part_rt[3])
    dict_PMMA[names[:-4]].append(part_rt[4])
    if date_sample >= date_kali_aug:
        dict_PA[names[:-4]].append(part_rt[5])
        dict_SBR[names[:-4]].append(part_rt[6])

    # Qualis
    dict_PE[names[:-4]].append(part_q[0])
    dict_PE[names[:-4]].append(part_q[1])
    dict_PE[names[:-4]].append(part_q[2])

    dict_PP[names[:-4]].append(part_q[3])
    dict_PP[names[:-4]].append(part_q[4])
    dict_PP[names[:-4]].append(part_q[5])

    dict_PS[names[:-4]].append(part_q[6])
    dict_PS[names[:-4]].append(part_q[7])
    dict_PS[names[:-4]].append(part_q[8])

    dict_PET[names[:-4]].append(part_q[9])
    dict_PET[names[:-4]].append(part_q[10])
    dict_PET[names[:-4]].append(part_q[11])

    dict_PMMA[names[:-4]].append(part_q[12])
    dict_PMMA[names[:-4]].append(part_q[13])
    dict_PMMA[names[:-4]].append(part_q[14])

    if date_sample >= date_kali_aug:
        dict_PA[names[:-4]].append(part_q[15])
        dict_PA[names[:-4]].append(part_q[16])
        dict_PA[names[:-4]].append(part_q[17])

        dict_SBR[names[:-4]].append(part_q[18])
        dict_SBR[names[:-4]].append(part_q[19])
        dict_SBR[names[:-4]].append(part_q[20])

    # Ratio
    dict_PE[names[:-4]].append(ratio)

# Erstellung dataframes und Bewertung PArameteranzahl

choices_PE = [gut, mittelgut, mittel, mittelschlecht, schlecht]

PE_final = dict(dict_PE)
PE_df = pd.DataFrame.from_dict(PE_final, orient='index',
                               columns=['RT', 'Qualifier_1', 'Qualifier_2', 'Qualifier_3', 'Ratio_55/81'])
condi_PE = [(PE_df.isnull().sum(axis=1) == 0),
            (PE_df.isnull().sum(axis=1) == 1),
            (PE_df.isnull().sum(axis=1) == 2),
            (PE_df.isnull().sum(axis=1) == 3),
            (PE_df.isnull().sum(axis=1) == 4)]
PE_df['Parameter_Anzahl'] = np.select(condi_PE, choices_PE, np.nan)
PE_df['Gesamt'] = PE_df.mean(axis=1)
PE_df['Gesamt'] = PE_df['Gesamt'].round(2)

choices_rest = [gut, mittelgut, mittel, schlecht]
dicts_polys = [dict_PP, dict_PS, dict_PET, dict_PMMA]
dicts_polys_new = [dict_PP, dict_PS, dict_PET, dict_PMMA, dict_PA, dict_SBR]
poly_gesamt = []

if not dict_PA == None:
    dicts_polys = dicts_polys_new
else: dicts_polys = dicts_polys

for dicts in dicts_polys:
    Poly_final = dict(dicts)
    Poly_final_df = pd.DataFrame.from_dict(Poly_final, orient='index', columns=['RT', 'Qualifier_1', 'Qualifier_2', 'Qualifier_3'])
    condi_poly = [(Poly_final_df.isnull().sum(axis=1) == 0),
        (Poly_final_df.isnull().sum(axis=1) == 1),
        (Poly_final_df.isnull().sum(axis=1) == 2),
        (Poly_final_df.isnull().sum(axis=1) == 3)]
    Poly_final_df['Parameter_Anzahl'] = np.select(condi_poly, choices_rest, np.nan)
    Poly_final_df['Gesamt'] = Poly_final_df.mean(axis=1)
    Poly_final_df['Gesamt'] = Poly_final_df['Gesamt'].round(2)
    poly_gesamt.append(Poly_final_df)

# MP-Berechnung
## Excel-Datei mit Kalibriergradeninfos einlesen
### Alle Infos sind in einer Excel zusammengefasst

data_june = pd.read_excel(r'Kalibrierungsdaten\NG_und_BG\NG_und_BG.xlsx', sheet_name='june_2020')
data_june.drop(['Berechneter Blindwert [µg]', 'Polymer'], axis=1, inplace=True)
data_june.rename(columns={'Nachweisgrenze xNG [µg]': 'NG', 'Bestimmungs-grenze xBG [µg]': 'BG'}, inplace=True)

data_nov = pd.read_excel(r'Kalibrierungsdaten\NG_und_BG\NG_und_BG.xlsx', sheet_name='nov_2020')
data_nov.drop(['Berechneter Blindwert [µg]', 'Polymer'], axis=1, inplace=True)
data_nov.rename(columns={'Nachweisgrenze xNG [µg]': 'NG', 'Bestimmungs-grenze xBG [µg]': 'BG'}, inplace=True)

data_mar = pd.read_excel(r'Kalibrierungsdaten\NG_und_BG\NG_und_BG.xlsx', sheet_name='mar_2021')
data_mar.drop(['Berechneter Blindwert [µg]', 'Polymer'], axis=1, inplace=True)
data_mar.rename(columns={'Nachweisgrenze xNG [µg]': 'NG', 'Bestimmungs-grenze xBG [µg]': 'BG'}, inplace=True)

data_aug = pd.read_excel(r'Kalibrierungsdaten\NG_und_BG\NG_und_BG.xlsx', sheet_name='aug_2021')
data_aug.drop(['Berechneter Blindwert [µg]', 'Polymer'], axis=1, inplace=True)
data_aug.rename(columns={'Nachweisgrenze xNG [µg]': 'NG', 'Bestimmungs-grenze xBG [µg]': 'BG'}, inplace=True)

# data_may = pd.read_excel(r'Kalibrierungsdaten\NG_und_BG\NG_und_BG.xlsx', sheet_name='may_2021_del')
# data_may.drop(['Berechneter Blindwert [µg]', 'Polymer'], axis=1, inplace=True)
# data_may.rename(columns={'Nachweisgrenze xNG [µg]': 'NG', 'Bestimmungs-grenze xBG [µg]': 'BG'}, inplace=True)

# Proben nochmal einlesen, diesmal mit BW-file
## BW sollte immer ganz oben stehen, also zu Not so umbenennen, dass BW oberste file ist
### Es können auch von mehreren Tagen gleichzeitig files eingelesen werden

open_samples = [open(f, 'r+') for f in filenames_samples]

sample_list = read_data(open_samples)

filenames_samples = [s.replace('Proben_Rohdaten\\', '') for s in filenames_samples]

# BW-Daten + Probendaten einlesen, BW-Korrektur und MP-Berechnung

mp_berechnung = sample_list
df_mp_list = []

for df, name in zip(mp_berechnung, filenames_samples):
    date = name[:6]
    date_sample = datetime.strptime(date, '%y%m%d')

    if date_sample < date_kali_nov:
         df_calibration = data_june
    elif date_kali_nov < date_sample < date_kali_mar:
         df_calibration = data_nov
    elif date_kali_mar < date_sample < date_kali_aug:
         df_calibration = data_mar
    else:
         df_calibration = data_aug

    if name[:-4].endswith('BW'):
        blanc_dict = dict(df)
        blanc_df = pd.DataFrame.from_dict(blanc_dict, orient='index',
                                          columns=['RT', 'RT_start', 'Response-BW', 'Quali1', 'Quali2', 'Quali3'])
        #print(blanc_df)
        blanc_df.reset_index(level=0, inplace=True)
        blanc_df.rename(columns={'index': 'Methoden-Name'}, inplace=True)
        y_istd = blanc_df.loc[1, 'Response-BW']
        blanc_df['BW_res'] = blanc_df['Response-BW'] / y_istd
        blanc_df['BW_res'] = blanc_df['BW_res'].replace(np.nan, 0.0)
        blanc_df = blanc_df.drop(['RT', 'RT_start', 'Response-BW', 'Quali1', 'Quali2', 'Quali3'], axis=1)
    elif 'BW' not in name[:-4]:
        mp_dict = dict(df)
        mp_df = pd.DataFrame.from_dict(mp_dict, orient='index',
                                       columns=['RT', 'RT_start', 'Response', 'Quali1', 'Quali2', 'Quali3'])
        mp_df.reset_index(level=0, inplace=True)
        mp_df.rename(columns={'index': 'Methoden-Name'}, inplace=True)
        y_istd = mp_df.loc[1, 'Response']

        mp_df['Res_korr'] = (mp_df['Response'] / y_istd) - blanc_df['BW_res']

        mp_df.drop(['RT', 'RT_start', 'Response', 'Quali1', 'Quali2', 'Quali3'], axis=1, inplace=True)

        df_mp_concat = pd.concat([mp_df, df_calibration], axis=0)
        df_mp_concat = df_mp_concat.groupby('Methoden-Name').sum()

        df_mp_concat['MP-Wert'] = (df_mp_concat['Res_korr'] - df_mp_concat['a*']) / df_mp_concat['b*']

        var_ab = - df_mp_concat['a*'] / df_mp_concat['b*']

        df_mp_concat['MP_final'] = df_mp_concat['MP-Wert'].where(
            (df_mp_concat['MP-Wert'] >= df_mp_concat['NG']) & (df_mp_concat['MP-Wert'] != var_ab), np.nan)
        df_mp_concat['MP_final'] = df_mp_concat['MP_final'].round(2)

        df_mp_concat.drop(['Res_korr', 'Verfahrens-standardab-weichung sx,0 [µg]', 'NG', 'BG', 'a*',
                           'b*', 'N (Anzahl Gesamtmessungen)', 'R2', 'Minimaler Wert [µg]', 'Maximaler Wert [µg]',
                           'MP-Wert'], axis=1, inplace=True)

        df_mp_list.append(df_mp_concat)

# Transponieren der Polymerwerte, die kalibriert wurden

final_list = []
for df_fin, name in zip(df_mp_list, filenames_samples):
    date = name[:6]
    date_sample = datetime.strptime(date, '%y%m%d')
    df_final = df_fin.transpose()
    if date_sample < date_kali_may:
        df_final.drop(['ISTD_d8_2_4_6_Triphenyl_1_hexen', 'ISTD_d8_2_4_Diphenyl_1_butene', 'ISTD_d8_Styrene',
                       'PET2_Methylbenzoat', 'PET4_Ethylbenzoat',
                       'PET6_Dimethyl_phthalat', 'PVC11_Benzene', 'PVC1_Mesitylene1',
                       'PVC3_Mesitylene3', 'PVC5_Biphenyl', 'PVC6_Acenaphthylene', 'PVC7_Anthracene', 'd8_Styrene'], axis=1, inplace=True)
    elif date_kali_may < date_sample < date_kali_aug:
        df_final.drop(['ISTD_d8_2_4_6_Triphenyl_1_hexen', 'ISTD_d8_2_4_Diphenyl_1_butene', 'ISTD_d8_Styrene',
                       'PET1_Acetophenon', 'PET4_Ethylbenzoat','Add1_DEHP', 'Add2_BBP', 'Add3_DOTP', 'PA1_caprolactam',
                       'PET6_Dimethyl_phthalat'], axis=1, inplace=True)
    else:
         df_final.drop(['ISTD_d8_2_4_6_Triphenyl_1_hexen', 'ISTD_d8_2_4_Diphenyl_1_butene', 'ISTD_d8_Styrene',
                       'Add1_DEHP', 'Add2_BBP', 'Add3_DOTP', 'PET1_Acetophenon',
                       'PET6_Dimethyl_phthalat'], axis=1, inplace=True)
    final_list.append(df_final)

merged_final = pd.concat(final_list)
merged_final.reset_index(drop=True, inplace=True)

#Hinzufügen Proben-ID

filename_list = []
for i in filenames_samples:
    if 'BW' not in i[:-4]:
        name = i[:-4]
        filename_list.append(name)

merged_final.insert(0, 'Proben-ID', filename_list)
merged_final = merged_final.set_index('Proben-ID')

# Splitten des gesamten Dataframes, um Werte des untersuchten Polymers in das jeweilige Excelsheet einzufügen

MP_PE_df = merged_final.iloc[:,0:12]
MP_PP_df = merged_final.iloc[:,18:24]
MP_PS_df = merged_final.iloc[:,24:27]
MP_PET_df = merged_final.iloc[:,12:17]
MP_PMMA_df = merged_final.iloc[:,17:18]

if not dict_PA == {}:
    MP_PE_df = merged_final.iloc[:, 1:13]
    MP_PP_df = merged_final.iloc[:, 20:26]
    MP_PS_df = merged_final.iloc[:, 26:29]
    MP_PET_df = merged_final.iloc[:, 13:19]
    MP_PMMA_df = merged_final.iloc[:, 19:20]
    MP_PA_df = merged_final.iloc[:, 0:1]
    MP_SBR_df = merged_final.iloc[:, 30:31]

# Hinzufügen Gesamtbewertung_1
PE_ges = PE_df.iloc[:,6:7]
PE_ges.rename(columns={'Gesamt':'Bewertung_1'}, inplace=True)
PP_ges = poly_gesamt[0].iloc[:,5:6]
PP_ges.rename(columns={'Gesamt':'Bewertung_1'}, inplace=True)
PS_ges = poly_gesamt[1].iloc[:,5:6]
PS_ges.rename(columns={'Gesamt':'Bewertung_1'}, inplace=True)
PET_ges = poly_gesamt[2].iloc[:,5:6]
PET_ges.rename(columns={'Gesamt':'Bewertung_1'}, inplace=True)
PMMA_ges = poly_gesamt[3].iloc[:,5:6]
PMMA_ges.rename(columns={'Gesamt':'Bewertung_1'}, inplace=True)

if not dict_PA == {}:
    PA_ges = poly_gesamt[4].iloc[:, 5:6]
    PA_ges.rename(columns={'Gesamt': 'Bewertung_1'}, inplace=True)
    SBR_ges = poly_gesamt[5].iloc[:, 5:6]
    SBR_ges.rename(columns={'Gesamt': 'Bewertung_1'}, inplace=True)


#_____________________________TGA-Daten hinzufügen____________________________________#
df_tga = run_tga()
df_tga.rename(columns={'Sample': 'Proben-ID'}, inplace=True)
df_tga = df_tga.set_index('Proben-ID')
df_tga = df_tga.round({'TGA Faktor': 3, 'Standardabweichung [%]': 3})

# Zusammenfügen von MP-Werten und Bewertung_1 um sie automatisch in MP_Zusammenfassung einzuschreiben

MP_PE_df2 = MP_PE_df
PE_liste = [MP_PE_df, MP_PE_df2, df_tga, PE_ges]
merged_PE = pd.concat(PE_liste, axis=1)

MP_PP_df2 = MP_PP_df
PP_liste = [MP_PP_df, MP_PP_df2, df_tga, PP_ges]
merged_PP = pd.concat(PP_liste, axis=1)

MP_PS_df2 = MP_PS_df
PS_liste = [MP_PS_df, MP_PS_df2, df_tga, PS_ges]
merged_PS = pd.concat(PS_liste, axis=1)

PET_liste = [MP_PET_df, df_tga, PET_ges]
merged_PET = pd.concat(PET_liste, axis=1)

PMMA_liste = [MP_PMMA_df, df_tga,  PMMA_ges]
merged_PMMA = pd.concat(PMMA_liste, axis=1)

if not dict_PA == {}:
    PA_liste = [MP_PA_df, df_tga, PA_ges]
    merged_PA = pd.concat(PA_liste, axis=1)

    SBR_liste = [MP_SBR_df, df_tga, SBR_ges]
    merged_SBR = pd.concat(SBR_liste, axis=1)


# Finale Eingabe in Zusammenfassungs-Excel

workbook = load_workbook('Final_Excels\\MP_calc_ohneMatrix_TGA.xlsx')
writer = pd.ExcelWriter('Final_Excels\\MP_calc_ohneMatrix_TGA.xlsx', engine='openpyxl')
writer.book = workbook
writer.sheets = {ws.title: ws for ws in workbook.worksheets}

merged_PE.to_excel(writer, sheet_name='PE', startcol=0, startrow=writer.sheets['PE'].max_row, index = True, header=False)
merged_PP.to_excel(writer, sheet_name='PP', startcol=0, startrow=writer.sheets['PP'].max_row, index = True, header=False)
merged_PS.to_excel(writer, sheet_name='PS', startcol=0, startrow=writer.sheets['PS'].max_row, index = True, header=False)
merged_PET.to_excel(writer, sheet_name='PET', startcol=0, startrow=writer.sheets['PET'].max_row, index = True, header=False)
merged_PMMA.to_excel(writer, sheet_name='PMMA', startcol=0, startrow=writer.sheets['PMMA'].max_row, index = True, header=False)
if not dict_PA == {}:
    merged_PA.to_excel(writer, sheet_name='PA', startcol=0, startrow=writer.sheets['PA'].max_row, index = True, header=False)
    merged_SBR.to_excel(writer, sheet_name='SBR', startcol=0, startrow=writer.sheets['SBR'].max_row, index = True, header=False)

PE_df.to_excel(writer, sheet_name='Bewertung_PE', startcol=0, startrow=writer.sheets['Bewertung_PE'].max_row, index = True, header=False)
poly_gesamt[0].to_excel(writer, sheet_name='Bewertung_PP', startcol=0, startrow=writer.sheets['Bewertung_PP'].max_row, index = True, header=False)

writer.save()
writer.close()

#save_excel('Final_Excels\MP_calc_ohneMatrix - Kopie.xlsx')

# Probenfiles müssen manuell in "done" verschoben werden

