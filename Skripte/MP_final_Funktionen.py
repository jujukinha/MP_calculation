from collections import defaultdict
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

# Funktion um detailed report einzulesen und in einem dictionary zu speichern

def read_data(open_data):
    file_list = []
    for original_files in (open_data):
        dict_raw = defaultdict(list)
        for line in original_files.readlines()[44:]:
            line = line.lstrip()
            if 'Compound' in line:
                name = str('\n' + line[16:])
                name = name.replace('-', '_')
                name = name.replace(',', '_')
            if 'NOT FOUND' in line:
                dict_raw[name[1:-1]].append(np.nan)
                dict_raw[name[1:-1]].append(np.nan)
                dict_raw[name[1:-1]].append(np.nan)
                dict_raw[name[1:-1]].append(np.nan)
                dict_raw[name[1:-1]].append(np.nan)
                dict_raw[name[1:-1]].append(np.nan)
            if 'Tgt' in line:
                dict_raw[name[1:-1]].append(float(line[35:40]))  # RT
                dict_raw[name[1:-1]].append(float(line[43:48]))  # RT_start
                dict_raw[name[1:-1]].append(float(line[50:61]))  # Response
            if 'Q1' in line:
                dict_raw[name[1:-1]].append(float(line[13:19]))  # Wert Qualifier1
            if 'Q2' in line:
                dict_raw[name[1:-1]].append(float(line[13:19]))  # Wert Qualifier2
            if 'Q3' in line:
                dict_raw[name[1:-1]].append(float(line[13:19]))  # Wert Qualifier3

        file_list.append(dict_raw)
        original_files.close()

    return file_list


def ratio_PE(dic_lists):
    df_ratio_PE = []
    for dicts in dic_lists:
        search_key = 'dien'
        neues_dict = dict(filter(lambda item: search_key in item[0], dicts.items()))
        neuesdict_df = pd.DataFrame.from_dict(neues_dict, orient='index',
                                              columns=['RT', 'RT_start', 'Response', 'Quali1', 'Quali2', 'Quali3'])
        neuesdict_df.drop(['RT', 'RT_start', 'Quali1', 'Quali2', 'Quali3'], axis=1, inplace=True)
        neuesdict_df.reset_index(level=0, inplace=True)
        neuesdict_df.rename(columns={'index': 'Proben-ID'}, inplace=True)
        new = neuesdict_df['Proben-ID'].str.split("_", n=1, expand=True)
        new = new.rename(columns={0: 'Name', 1: 'specific_name'})
        # print(new)
        neuesdict_df['Name'] = new['Name']
        neuesdict_df = neuesdict_df.groupby('Name')[['Response']].apply(lambda x: x.max() / x.min())

        neuesdict_df = neuesdict_df.drop(['SBR1'])

        neuesdict_df = neuesdict_df.round(1)

        df_ratio_PE.append(neuesdict_df)

    return df_ratio_PE

def ratio_PE_mitMatrix(dic_lists):
    df_ratio_PE = []
    for dicts in dic_lists:
        search_key = 'dien'
        neues_dict = dict(filter(lambda item: search_key in item[0], dicts.items()))

        neuesdict_df = pd.DataFrame.from_dict(neues_dict, orient='index',
                                              columns=['RT', 'RT_start', 'Response', 'Quali1', 'Quali2', 'Quali3'])
        neuesdict_df.drop(['RT', 'RT_start', 'Quali1', 'Quali2', 'Quali3'], axis=1, inplace=True)
        neuesdict_df.reset_index(level=0, inplace=True)
        neuesdict_df.rename(columns={'index': 'Proben-ID'}, inplace=True)
        new = neuesdict_df['Proben-ID'].str.split("_", n=1, expand=True)
        new = new.rename(columns={0: 'Name', 1: 'specific_name'})
        # print(new)
        neuesdict_df['Name'] = new['Name']
        neuesdict_df = neuesdict_df.groupby('Name')[['Response']].apply(lambda x: x.max() / x.min())
        neuesdict_df = neuesdict_df.round(1)
        neuesdict_df = neuesdict_df.drop(['PE1', 'PE2', 'PE6', 'SBR1'])

        df_ratio_PE.append(neuesdict_df)

    return df_ratio_PE

def save_excel(excelpath, new_poly=None):

    workbook = load_workbook(excelpath)
    writer = pd.ExcelWriter(excelpath, engine='openpyxl')
    writer.book = workbook
    writer.sheets = {ws.title: ws for ws in workbook.worksheets}

    merged_PE.to_excel(writer, sheet_name='PE', startcol=0, startrow=writer.sheets['PE'].max_row, index = True, header=False)
    merged_PP.to_excel(writer, sheet_name='PP', startcol=0, startrow=writer.sheets['PP'].max_row, index = True, header=False)
    merged_PS.to_excel(writer, sheet_name='PS', startcol=0, startrow=writer.sheets['PS'].max_row, index = True, header=False)
    merged_PET.to_excel(writer, sheet_name='PET', startcol=0, startrow=writer.sheets['PET'].max_row, index = True, header=False)
    merged_PMMA.to_excel(writer, sheet_name='PMMA', startcol=0, startrow=writer.sheets['PMMA'].max_row, index = True, header=False)
    if new_poly is not None:
        merged_PA.to_excel(writer, sheet_name='PA', startcol=0, startrow=writer.sheets['PA'].max_row, index = True, header=False)

    PE_df.to_excel(writer, sheet_name='Bewertung_PE', startcol=0, startrow=writer.sheets['Bewertung_PE'].max_row, index = True, header=False)
    poly_gesamt[0].to_excel(writer, sheet_name='Bewertung_PP', startcol=0, startrow=writer.sheets['Bewertung_PP'].max_row, index = True, header=False)

    writer.save()
    writer.close()

    return writer